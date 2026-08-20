import os
import glob
from typing import List, Dict, Any
import pymupdf
import docx
import openpyxl
from langchain_text_splitters import RecursiveCharacterTextSplitter

def process_pdf(file_path: str, department: str, chunk_size: int, chunk_overlap: int) -> List[Dict[str, Any]]:
    """Extracts and chunks text from a PDF."""
    doc_name = os.path.basename(file_path)
    pages_data = []
    try:
        doc = pymupdf.open(file_path)
        for page_num, page in enumerate(doc):
            text = page.get_text()
            if text.strip():
                pages_data.append({
                    "page_number": page_num + 1,
                    "text": text.strip()
                })
        doc.close()
    except Exception as e:
        print(f"Error reading PDF {file_path}: {e}")
        return []

    return chunk_pages_data(pages_data, doc_name, department, chunk_size, chunk_overlap)

def process_docx(file_path: str, department: str, chunk_size: int, chunk_overlap: int) -> List[Dict[str, Any]]:
    """Extracts and chunks text from a DOCX file."""
    doc_name = os.path.basename(file_path)
    pages_data = []
    try:
        doc = docx.Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        if text.strip():
            pages_data.append({
                "page_number": 1,  # Treat docx as a single long page
                "text": text
            })
    except Exception as e:
        print(f"Error reading DOCX {file_path}: {e}")
        return []

    return chunk_pages_data(pages_data, doc_name, department, chunk_size, chunk_overlap)

def chunk_pages_data(pages_data: List[Dict[str, Any]], doc_name: str, department: str, chunk_size: int, chunk_overlap: int) -> List[Dict[str, Any]]:
    """Helper function to chunk text using Langchain's text splitter."""
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        length_function=len,
        is_separator_regex=False,
    )
    chunks = []
    for page in pages_data:
        page_chunks = text_splitter.split_text(page["text"])
        for i, chunk_text in enumerate(page_chunks):
            chunk_id = f"{doc_name}_{department}_p{page['page_number']}_c{i}"
            chunks.append({
                "chunk_id": chunk_id,
                "document_name": doc_name,
                "department": department,
                "page_number": page["page_number"],
                "text": chunk_text
            })
    return chunks

def process_xlsx(file_path: str, department: str) -> List[Dict[str, Any]]:
    """Extracts text from an XLSX file, chunking by rows to preserve row range metadata."""
    doc_name = os.path.basename(file_path)
    chunks = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        rows_per_chunk = 15
        
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            current_chunk_rows = []
            start_row = 1
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Convert row to string, skip entirely empty rows
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip() and row_text.strip() != " | " * (len(row)-1):
                    current_chunk_rows.append(row_text)
                    
                if len(current_chunk_rows) >= rows_per_chunk:
                    end_row = row_idx
                    chunk_id = f"{doc_name}_{department}_{sheetname}_r{start_row}-{end_row}"
                    chunks.append({
                        "chunk_id": chunk_id,
                        "document_name": doc_name,
                        "department": department,
                        "page_number": f"Sheet:{sheetname}",
                        "sheet_name": sheetname,
                        "row_range": f"{start_row}-{end_row}",
                        "text": "\n".join(current_chunk_rows)
                    })
                    current_chunk_rows = []
                    start_row = row_idx + 1
                    
            if current_chunk_rows:
                end_row = sheet.max_row
                chunk_id = f"{doc_name}_{department}_{sheetname}_r{start_row}-{end_row}"
                chunks.append({
                    "chunk_id": chunk_id,
                    "document_name": doc_name,
                    "department": department,
                    "page_number": f"Sheet:{sheetname}",
                    "sheet_name": sheetname,
                    "row_range": f"{start_row}-{end_row}",
                    "text": "\n".join(current_chunk_rows)
                })
    except Exception as e:
        print(f"Error reading XLSX {file_path}: {e}")
    return chunks

def process_document(file_path: str, department: str, chunk_size: int = 1000, chunk_overlap: int = 200) -> List[Dict[str, Any]]:
    """
    Processes a document (PDF, DOCX, or XLSX) based on its extension.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return process_pdf(file_path, department, chunk_size, chunk_overlap)
    elif ext == ".docx":
        return process_docx(file_path, department, chunk_size, chunk_overlap)
    elif ext == ".xlsx":
        return process_xlsx(file_path, department)
    else:
        print(f"Unsupported file format: {ext}")
        return []

def ingest_all_departments(data_dir: str, chunk_size: int = 1000, chunk_overlap: int = 200) -> List[Dict[str, Any]]:
    """Reads all supported files from department folders."""
    all_chunks = []
    departments = ["Finance", "HR", "IT", "Legal", "Sales"]
    
    for dept in departments:
        dept_path = os.path.join(data_dir, dept)
        if not os.path.exists(dept_path):
            continue
            
        files = glob.glob(os.path.join(dept_path, "*.*"))
        for file_path in files:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in [".pdf", ".docx", ".xlsx"]:
                doc_chunks = process_document(file_path, dept, chunk_size, chunk_overlap)
                all_chunks.extend(doc_chunks)
            
    return all_chunks

if __name__ == "__main__":
    # Example usage
    data_directory = os.path.join(os.path.dirname(__file__), "..", "..", "data")
    chunks = ingest_all_departments(data_directory)
    print(f"Total chunks generated: {len(chunks)}")
    if chunks:
        print("Sample chunk:")
        print(chunks[0])
